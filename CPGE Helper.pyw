from tkinter import *
from random import choice
from PIL import Image, ImageTk
from tkinter import filedialog
import openpyxl
from re import compile
from datetime import datetime
from os.path import basename
from sys import exit as leave_all
import os
from subprocess import run
import webbrowser
import csv
from tkinter import messagebox

#############
##FONCTIONS##
#############

def on_mousewheel(event):
    """
    ROLE: Scrollbar vérticale de canvas
    """
    global canvas
    canvas.yview_scroll(-1 * int(event.delta / 120), "units")
    
def on_mousewheel_horizontal(event):
    """
    ROLE: Scrollbar horizontale de canvas
    """
    global canvas
    canvas.xview_scroll(-1 * int(event.delta / 120), "units")

def leave_prgm():
    """
    ROLE: Quitte totalement le programme
    """
    leave_all()

def back_button_func():
    """
    ROLE: Retour à l'affichage initiale dans colonne_3 (hors du menu)
    """
    clear(colonne_3)
    Label(colonne_3, text="CPGE Helper", bg="#020E26", fg="white", font="Verdana 46 bold").pack(expand=True, anchor="center")

def reload_tas():
    """
    ROLE: Interface pour le tirage au sort
    """
    global canvas
    def suppr_grp(path_grp_file):
        """
        ROLE: Supprime un groupe en supprimant son fichier et reload l'interface TAS (action du bouton X)
        PARAMETRE: path_grp_file -> str
        """
        f = open(path_grp_file, 'r', encoding="utf-8")
        title = f.readline()
        f.close()
        if messagebox.askyesno('Question', "Voulez-vous supprimer le groupe {}".format(title)):
            os.remove(path_grp_file)
            reload_tas()
        
    def edit_grp(path_grp_file):
        """
        ROLE: Ouverture d'un edituer de texte pour modifier un groupe et reload l'interface TAS (action du bouton Modifier)
        PARAMETRE: path_grp_file -> str
        """
        try: #macOS
            run(["open", "-e", path_grp_file])
        except:
            try: #windows 
                run(["notepad.exe", path_grp_file])
            except:
                try: #linux
                    run(["xdg-open", path_grp_file])
                except:
                    print("ERREUR - Ouverture de l'éditeur de texte impossible !")
                    reload_main()
        reload_tas()
    
    def add_grp():
        """
        ROLE: Ouvre un éditeur de texte pour compléter les informations d'un nouveau groupe et reload l'interface TAS (action du bouton Ajouter)
        """
        f = open("data/group_config.txt", "r")
        n = int(f.readline()) #nb de groupe pour le nom du fichier
        f.close()
        
        while (os.path.exists("data/groupes/grp{}.txt".format(n))):
            n += 1
        line_to_write = ["Nom du groupe\n", "Prénom|0\n", "Prénom|0\n", "Prénom|0\n"] 
        f = open("data/groupes/grp{}.txt".format(n), "w", encoding="utf-8")
        for el in line_to_write:
            f.write(el)
        f.close()
        try: #macOS
            run(["open", "-e", "data/groupes/grp{}.txt".format(n)])
        except:
            try: #windows 
                run(["notepad.exe", "data/groupes/grp{}.txt".format(n)])
            except:
                try: #linux
                    run(["xdg-open", "data/groupes/grp{}.txt".format(n)])
                except:
                    print("ERREUR - Ouverture de l'éditeur de texte impossible !")
                    reload_main()
        f = open("data/group_config.txt", "w")
        f.write(str(n+1) + "\n")
        f.close()
        reload_tas()
        
    def choose_grp(path_grp_file, name_grp):
        """
        ROLE: Interface lorsque un groupe est choisi pour un tirage au sort (action du bouton Sélectionner)
        PARAMETRE: path_grp_file -> str ; name_grp -> str
        """
        global photo_back, photo_add
        def make_tirage(path_grp_file, n=1):
            """
            ROLE: Effectue un tirage et actualise la ligne correspondante au membre dans le fichier du groupe (action du bouton Lancer un tirage)
            PARAMETRES: path_grp_file -> str, n -> int
            """
            def remove_pick(content_of_file_list, index, path):
                """
                ROlE: Remet la ligne correspondante au membre du groupe tiré à son état précédent dans le fichier du groupe (action du bouton Annuler)
                PARAMETRES: content_of_file_list -> list ; index -> int ; path -> str
                """
                for pack in selected_box.pack_slaves():
                        pack.destroy()
                content_of_file_list[index] = content_of_file_list[index].split("|")[0] + "|" + str(int(content_of_file_list[index].split("|")[1][:-1]) - 1) + "\n"
                f = open(path, "w", encoding="utf-8")
                for el in content_of_file_list:
                    f.write(el)
                f.close()
            
            for i in range(n):
                fi = open(path_grp_file, "r", encoding="utf-8")
                l = fi.readlines()
                fi.close() #ca ne récupère pas bien les int apres les prénoms a cause des nombres à + de 1 chiffres (jsuis con)
                try:
                    m = min([int(el.strip().split("|")[1]) for el in l[1:]])
                    check_list = []
                    for i in range(1, len(l)):
                        actual_line = l[i].strip().split("|")
                        if int(actual_line[1]) == m:
                            check_list.append((i, actual_line[0], m+1)) # (indice, "prénom")
                    name_selected = (choice(check_list))
                    l[name_selected[0]] = name_selected[1] + "|" + str(m+1) + "\n" #"prénom + (min+1)"
                    fi = open(path_grp_file, "w", encoding="utf-8")
                    for el in l:
                        fi.write(el)
                    fi.close()
                    
                    for pack in selected_box.pack_slaves():
                        pack.destroy()
                    
                    Label(selected_box, text=name_selected[1], font="Verdana 40 bold", fg="white", bg="#152C5E", height=1, highlightthickness=2, highlightbackground="black").pack(fill="x", ipadx=15)
                    Label(selected_box, text= "Passage {}".format(name_selected[2]), font="Verdana 15 bold", bg="#344E89", fg="white", height=1, highlightthickness=2, highlightbackground="black").pack(fill="x", ipady=3)
                    Button(selected_box, text="Annuler", font="Verdana 15 bold", fg="white", bg="red", bd=2, relief="solid", height=1, command=lambda: remove_pick(l, name_selected[0], path_grp_file)).pack(fill="x")
                    
                except:
                    print("""ERREUR - Fichier de groupe invalide\nNom du groupe: "{}"\nAdresse du fichier: "{}\"""".format(name_grp, path_grp_file))
                    reload_tas()
                    
        clear(colonne_3)
        
        Label(colonne_3, text="Tirage au sort dans " + name_grp, font="Verdana 20 bold", fg="white", bg="#020E26").pack(pady=25)
        
        selected_box = Frame(colonne_3, bg="#020E26")
        selected_box.pack(pady=15, expand=True)
        
        Button(colonne_3, text="Lancer un tirage", cursor="hand2", font="Verdana 20 bold", fg="white", activeforeground="white", bg="#020E26", activebackground="#020E26", bd=3, command=lambda: make_tirage(path_grp_file)).pack(pady=15)

        image_back = Image.open("data/img/return_button.png")
        photo_back = ImageTk.PhotoImage(image_back)
        Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=reload_tas).pack(pady=15)
    
    global photo_back, canvas
    
    clear(colonne_3)
    
    Label(colonne_3, text="Tirage au sort", font="Verdana 25 bold", fg="white", bg="#020E26").pack(pady=15)
    
    grp_valid = False #check si TOUS les fichiers étaient vide ou pas (pour afficher le msg "aucun groupe")
    grp_file_list = os.listdir("data/groupes") #tout les fichiers contenant les groupes
    grp_box = Frame(colonne_3, bg="#020E26")  #cadre grp_box dans la colonne principale
    if grp_file_list:
        canvas = Canvas(grp_box, bg="#020E26", bd=0, highlightthickness=0, width=700, height=250)  
        canvas.pack(side="left", fill="both", expand=True)  
        scrollbar = Scrollbar(grp_box, orient=HORIZONTAL, bg="#020E26", troughcolor="#020E26", activebackground="#020E26", highlightthickness=0)
        scrollbar.pack(side="bottom", fill="x") 

        canvas.configure(xscrollcommand=scrollbar.set)
        fen.bind("<MouseWheel>", on_mousewheel_horizontal)
                
        grp_container = Frame(canvas, bg="#020E26")  
        canvas.create_window((0, 0), window=grp_container, anchor="nw")  
        
        for i, el in enumerate(os.listdir("data/groupes")):
            f = open("data/groupes/" + el, "r", encoding="utf-8")
            l = f.readlines()
            f.close()
            
            if l == []: #le fichier est vide
                os.remove("data/groupes/" + el)
            else: #le fichier est pas vide (valide)
                grp_valid = True
                one_grp_box = Frame(grp_container, bg="#344E89", highlightthickness=0, highlightbackground="black")
                one_grp_box.grid(row=0, column=i, padx=15)
                
                title_box = Frame(one_grp_box, bg="#152C5E", highlightthickness=2, highlightbackground="black")
                title_box.pack(expand=True, fill="x")
                Label(title_box, text=l[0][:-1], font="Verdana 15 bold", bg="#152C5E", fg="white").pack(pady=10)
                
                member_box = Frame(one_grp_box, bg="#344E89", highlightbackground="black", highlightthickness=2)
                member_box.pack(expand=True, fill="x")
                if len(l) > 3:
                    max_name = 4
                else:
                    max_name = len(l)
                Label(member_box, text="Membres ({}):".format(len(l)-1), font="Verdana 12 bold", bg="#344E89", fg="white").pack()
                for j in range(1, max_name):
                    Label(member_box, text=l[j].strip().split("|")[0], font="Verdana 12", bg="#344E89", fg="white").pack()
                for j in range(1, 5 - max_name):
                    Label(member_box, text="", font="Verdana 12", bg="#344E89", fg="white").pack()
                
                edit_suppr_box = Frame(one_grp_box, bg="#344E89")
                edit_suppr_box.pack(fill="x", expand=True)
                edit_suppr_box.grid_columnconfigure(0, weight=1)
                edit_suppr_box.grid_columnconfigure(1, weight=1)

                btn1 = Button(edit_suppr_box, text="Modifier", relief=SOLID, bd=2, font="Verdana 15 bold", bg="#152C5E", fg="white", command=lambda x=el: edit_grp("data/groupes/" + x))
                btn1.grid(row=0, column=0, sticky="ew")  # Utilisation de "ew" pour remplir horizontalement

                btn2 = Button(edit_suppr_box, text="X", relief=SOLID, bd=2, font="Verdana 15 bold", bg="red", fg="white", command=lambda x = el: suppr_grp("data/groupes/" + x))
                btn2.grid(row=0, column=1, sticky="ew")  # Utilisation de "ew" pour remplir horizontalement

                btn3 = Button(one_grp_box, text="Sélectionner", relief="solid", bd=2, font="Verdana 15 bold", bg="#344E89", fg="white", command=lambda x=el, y=l[0][:-1]: choose_grp("data/groupes/" + x, y))
                btn3.pack(fill="x")

        canvas.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))
                
    if not grp_valid:
        Label(colonne_3, text="Aucun groupe", font="Verdana 15 bold", fg="white", bg="#020E26").pack(pady=20)
    
    grp_box.pack(expand=True)
    
    Button(colonne_3, text="Ajouter", font="Verdana 18 bold", bg="#020E26", fg="white", bd=4, highlightthickness=5, activebackground="#344E89", command=add_grp).pack()
    
    image_back = Image.open("data/img/return_button.png")
    photo_back = ImageTk.PhotoImage(image_back)
    Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=back_button_func).pack(pady=15)

def create_colloscope_prgm(title, nb_grp, sem_start, sem_end, dic_colleurs, dic_matieres):
    """
    ROLE: Génère le fichier CSV du colloscope à importer dans e-colle
    PARAMETRE: 
        - title -> str
        - nb_grp -> int
        - sem_start -> int
        - sem_end -> int
        - dic_colleurs -> dic[str] = dic[str] = [(str, str, str)]
        - dic_matieres -> dic[str] = (int, int, [int, int])
    RETURN: None
    """
    def get_good_frequence_list(l_initiale, l_finale):
        """
        ROLE: Renvoie la plus petite liste de groupes dont la somme des éléments est égal à 1
        PARAMETRES: l_initiale -> list ; l_finale -> list
        RETURN: list
        """
        #using backtracking
        return l_finale
    
    clear(colonne_3)
    
    Label(colonne_3, text="Créateur de colloscope", font="Verdana 25 bold", fg="white", bg="#020E26").pack(pady=15)
    
    Label(colonne_3, text="Création du colloscope en cours...", font="Verdana 13", fg="white", bg="#020E26").pack()
    
    col_into_list = [["Matière", "Nom", "Prénom", "Créneau", "Salle"]] #liste de liste (tableau lignes x colonnes)
    order_grp_list = [i for i in range(1, nb_grp+1)] #déacalage a chaque semaine
    
    for semaine in range(sem_start, sem_end+1): #Création de la premiere ligne
        col_into_list[0].append(",S{}".format(semaine))
    
    for matieres in dic_colleurs.keys(): #Création de la colonne 1, 2, 3 et 4
        for creneaux in dic_colleurs[matieres]:
            for colleurs in dic_colleurs[matieres][creneaux]:
                col_into_list.append([matieres, colleurs[0], colleurs[1], creneaux])
    print(col_into_list)
                
    #faire une liste décroissante de la somme égale à 1 des fréquences des matières
    #Permet de trouver les matières "liées" (exemple : anglais physique qui s'alterne)
            
            
def create_colloscope_gui(title=None, nb_grp=None, sem_start=None, sem_end=None, dic_colleurs=None, dic_matieres=None):
    """
    ROLE: Récupère grâce à l'interface les informations pour la création du khôlloscope
    PARAMETRE: 
        - title -> None | str
        - nb_grp -> None | int
        - sem_start -> None | int
        - sem_end -> None | int
        - dic_colleurs -> None | dic[str] = dic[str] = [(str, str, str)]
        - dic_matieres -> None | dic[str] = (int, int, [int, int])
    RETURN: None
    """
    def validate_all():
        """
        ROLE: Vérifie si toute les informations nécessaire à la création d'un khôlloscope ont été renseignée puis active la création (create_colloscope_prgm)
        """
        if title and nb_grp and sem_start and sem_end and dic_colleurs and dic_matieres:
            dir_save_path = filedialog.askdirectory()
            if dir_save_path:
                print("\nTitre :")
                print(title)
                print("\nNb_grp :")
                print(nb_grp)
                print("\nsem_start :")
                print(sem_start)
                print("\nsem_end :")
                print(sem_end)
                print("\ndic_colleurs :")
                print(dic_colleurs)
                print("\ndic_matieres :")
                print(dic_matieres)
                create_colloscope_prgm(title, nb_grp, sem_start, sem_end, dic_colleurs, dic_matieres)
        else:
            print("Info invalide /!\\")
    
    def config_title():
        """
        ROLE: Interface pour configurer le titre du khôlloscope en cours de création
        """
        def validate_title(content):
            """
            ROLE: Vérifie la compatibilité du titre entré
            PARAMATRE: content -> str
            """
            def test_title(content):
                """
                ROLE: Vérifie si les caractères interdits sont présents dans content
                PARAMETRE: content -> str
                """
                interdit = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
                for c in interdit:
                    if c in content:
                        return False
                return True

            if test_title(content) and len(content) > 0:
                create_colloscope_gui(content, nb_grp, sem_start, sem_end, dic_colleurs, dic_matieres)
        
        global photo_back
        
        clear(colonne_3)

        Label(colonne_3, text="Créateur de colloscope", font="Verdana 25 bold", fg="white", bg="#020E26").pack(pady=15)

        t_box = Frame(colonne_3, bg="#020E26")
        t_box.pack(expand=True)
        
        Label(t_box, text="Titre :", font="Verdana 20 bold", fg="white", bg="#020E26").grid(row=0, column=0)
        titre_entry = Entry(t_box, font="Verdana 20 bold")
        titre_entry.grid(row=0, column=1)
        
        Button(colonne_3, text="Valider", cursor="hand2", font="Verdana 20 bold", fg="white", activeforeground="white", bg="#020E26", activebackground="#020E26", bd=3, command=lambda: validate_title(titre_entry.get())).pack(pady=5)
    
        image_back = Image.open("data/img/return_button.png")
        photo_back = ImageTk.PhotoImage(image_back)
        Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=lambda: create_colloscope_gui(title, nb_grp, sem_start, sem_end, dic_colleurs, dic_matieres)).pack(pady=15)
    
    def config_matiere():
        """
        ROLE: Interface pour configurer les matières du khôlloscope en cours de création (fréquences)
        """
        def next_config_mat():
            """
            ROLE: Interface pour l'affichage de la deuxième page de configuration des matières du khôlloscope en cours de création (groupes)
            """
            def on_checkbox_toggle(label, indice):
                """
                ROLE: Actualise les boutons de sélections des groupes
                PARAMETRES: label -> tk.Label ; indice -> int
                """
                if label["bg"] == "red":
                    label.config(bg="green")
                    values[matieres[indice]][2].append(int(label["text"]))
                else:
                    label.config(bg="red")
                    values[matieres[indice]][2].remove(int(label["text"]))
                
            values = {}
            for el in spinbox_values:
                values[el[0]] = (int(el[1].get()), int(el[2].get()), [])
            
            #Amélioration future : transfo. de la matière en Button pour cocher toute la ligne des groupes
            for child in m_box.winfo_children():
                if child.grid_info()['column'] == 1:
                    child.destroy()
            finish_button.config(text="Valider", command=lambda: create_colloscope_gui(title, nb_grp, sem_start, sem_end, dic_colleurs, values))
            
            Label(m_box, text="Groupes", font="Verdana 13 bold", bg="#152C5E", fg="white", width=15, highlightbackground="black", highlightthickness=1).grid(row=0, column=1, sticky="ew", ipady=3)
            for i in range(len(matieres)):
                grp_box = Frame(m_box, bg="#344E89")
                grp_box.grid(row=i+1, column=1, sticky="ew")
                
                for c in range(nb_grp):
                    checkbox_label = Label(grp_box, text=str(c+1), font="Verdana 10", fg="white", bg="red", highlightthickness=1, highlightbackground="black")
                    checkbox_label.grid(row=0, column=c, sticky="ns")
                    checkbox_label.bind("<Button-1>", lambda event, label=checkbox_label, ind=i: on_checkbox_toggle(label, ind))
        
        global photo_back
        
        clear(colonne_3)
        
        Label(colonne_3, text="Créateur de colloscope", font="Verdana 25 bold", fg="white", bg="#020E26").pack(pady=15)        
        
        m_box = Frame(colonne_3, bg="#020E26")
        m_box.pack(expand=True)
        
        if not dic_colleurs or not nb_grp:
            Label(m_box, text="Veuillez d'abord configurer les colleurs et les groupes !", font="Verdana 12", bg="#020E26", fg="white").pack()
        else:
            matieres = [k for k in dic_colleurs.keys()]
            
            Label(m_box, text="Matières", font="Verdana 13 bold", bg="#152C5E", fg="white", width=15, highlightbackground="black", highlightthickness=1).grid(row=0, column=0, sticky="ew", ipady=3)
            Label(m_box, text="Fréquences", font="Verdana 13 bold", bg="#152C5E", fg="white", width=15, highlightbackground="black", highlightthickness=1).grid(row=0, column=1, sticky="ew", ipady=3)
            
            spinbox_values = []
            for i, el in enumerate(matieres):
                Label(m_box, text=el, font="Verdana 13", bg="#344E89", fg="white", width=15, highlightbackground="black", highlightthickness=1).grid(row=i+1, column=0, sticky="ew")
                
                freq_box = Frame(m_box, bg="#344E89", highlightbackground="black", highlightthickness=1)
                freq_box.grid(row=i+1, column=1, sticky="ew")
                spin1 = Spinbox(freq_box, font="Verdana 13", from_=1, to=10, width=3)
                spin1.grid(row=0, column=0)
                Label(freq_box, text="colles /", font="Verdana 13", bg="#344E89", fg="white").grid(row=0, column=1)
                spin2 = Spinbox(freq_box, font="Verdana 13", from_=1, to=4, width=3)
                spin2.grid(row=0, column=2)
                Label(freq_box, text="semaines", font="Verdana 13", bg="#344E89", fg="white").grid(row=0, column=3)
                spinbox_values.append((el ,spin1, spin2))
                
            finish_button = Button(colonne_3, text="Suivant", cursor="hand2", font="Verdana 20 bold", fg="white", activeforeground="white", bg="#020E26", activebackground="#020E26", bd=3, command=next_config_mat)
            finish_button.pack(pady=5)

        image_back = Image.open("data/img/return_button.png")
        photo_back = ImageTk.PhotoImage(image_back)
        Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=lambda: create_colloscope_gui(title, nb_grp, sem_start, sem_end, dic_colleurs, dic_matieres)).pack(pady=15)

    def config_grp():  
        """
        ROLE: Verifie si l'entrée de nb_grp est correcte et la valide
        """
        global photo_back
        
        try:
            nb_grp_temp = int(grp_spin.get())
        except:
            pass
        else:
            create_colloscope_gui(title, nb_grp_temp, sem_start, sem_end, dic_colleurs, dic_matieres)
    
    def config_sem():
        """
        ROLE: Verifie si les entrée pour les semaines sont correctent et les valident
        """
        try:
            x = int(sem_spin1.get())
            y = int(sem_spin2.get())
        except:
            pass
        else:
            if x < y:
                create_colloscope_gui(title, nb_grp, x, y, dic_colleurs, dic_matieres)
    
    def config_colleurs():
        """
        ROLE: Interface pour l'importation du fichier contenant les colleurs pour le khôlloscope en cours de création
        """
        def validate_import_file(dic):
            """
            ROLE: Valide le fichier selectionner (action du boutton valider)
            """
            if dic:
                create_colloscope_gui(title, nb_grp, sem_start, sem_end, dic, dic_matieres)
        
        def import_colleurs_file():
            """
            ROLE: Récupère les données d'un fichier sélectionné par l'utilisateur (fichier csv colleurs, voir documentation)
            """
            global m_dir
            
            file_path = filedialog.askopenfilename(title="Ouvrir un fichier", filetypes=(("Fichiers CSV", "*.csv"), ("CSV", "*.csv")))
            try:
                with open(file_path, "r", encoding="utf-8-sig") as f:
                    l = csv.reader(f)
                    lines_extracted = list(l)
                    
                lines = [el[0].split(";") for el in lines_extracted]
                temp_dir = {} #dico selon les matières
                for l in lines:
                    if l[0] in temp_dir:
                        if l[1] in temp_dir[l[0]]:
                            temp_dir[l[0]][l[1]].append((l[2], l[3], l[4]))
                        else:
                            temp_dir[l[0]][l[1]] = [(l[2], l[3], l[4])] 
                    else:
                        temp_dir[l[0]] = {l[1]: [(l[2], l[3], l[4])]}
                
                clear(config_colleurs_box)
                Button(config_colleurs_box, width=25 ,text="Changer de fichier colleurs", font="Verdana 19 bold", fg="white", activeforeground="white", bg="#020E26", activebackground="#020E26", bd=3, command=import_colleurs_file).pack()
                Label(config_colleurs_box, text="Fichier : {}".format(basename(file_path)), font="Verdana 15", bg="#020E26", fg="white").pack(pady=15)
                m_dir = temp_dir
            except:
                print("IMPORT FILE - Erreur de lecture du fichier")
        global photo_back, m_dir
        
        clear(colonne_3)
        
        m_dir = None
        
        Label(colonne_3, text="Créateur de colloscope", font="Verdana 25 bold", fg="white", bg="#020E26").pack(pady=15)
        config_colleurs_box = Frame(colonne_3, bg="#020E26")
        config_colleurs_box.pack(expand=True)
        indication_box = Frame(config_colleurs_box, bg="#020E26")
        indication_box.pack(pady=15)
        Label(indication_box, text="Disposition :", font="Verdana 10", fg="white", bg="#020E26").pack()
        Label(indication_box, text="Matière | Créneaux | Nom | Prénom", font="Verdana 10", fg="white", bg="#020E26").pack()
        Label(indication_box, text="Créneaux : 2 premières lettres du jours suivi de l'heure (ex: Je 13h00)", font="Verdana 10", fg="white", bg="#020E26").pack()
        Button(config_colleurs_box, width=25 ,text="Importer le fichier colleurs", font="Verdana 19 bold", fg="white", activeforeground="white", bg="#020E26", activebackground="#020E26", bd=3, command=import_colleurs_file).pack()
        Button(colonne_3, text="Valider", cursor="hand2", font="Verdana 20 bold", fg="white", activeforeground="white", bg="#020E26", activebackground="#020E26", bd=3, command=lambda: validate_import_file(m_dir)).pack(pady=5)
    
        image_back = Image.open("data/img/return_button.png")
        photo_back = ImageTk.PhotoImage(image_back)
        Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=lambda: create_colloscope_gui(title, nb_grp, sem_start, sem_end, dic_colleurs, dic_matieres)).pack(pady=15)
    
    global photo_back
    
    clear(colonne_3)
    
    Label(colonne_3, text="Créateur de colloscope", font="Verdana 25 bold", fg="white", bg="#020E26").pack(pady=15)
    
    all_config_box = Frame(colonne_3, bg="#020E26")
    all_config_box.pack(expand=True)

    Label(all_config_box, text="Catégories", font="Verdana 15", fg="white", bg="#152C5E", highlightbackground="black", highlightthickness=1).grid(row=0, column=0, sticky="ew")
    Label(all_config_box, text="Actions", font="Verdana 15", fg="white", bg="#152C5E", highlightbackground="black", highlightthickness=1).grid(row=0, column=1, sticky="ew")
    
    if not title:
        Label(all_config_box, text="Titre :", font="Verdana 15", fg="white", bg="#344E89", highlightbackground="black", highlightthickness=1).grid(row=1, column=0, sticky="nsew")
        Button(all_config_box, text="config.", font="Verdana 15", fg="white", bg="#020E26", command=config_title).grid(row=1, column=1, sticky="nsew")
    else:
        Label(all_config_box, text="Titre : \"{}\"".format(title), font="Verdana 15", fg="white", bg="#344E89", highlightbackground="black", highlightthickness=1).grid(row=1, column=0, sticky="nsew")
        Label(all_config_box, text="Fait !", font="Verdana 15", fg="white", bg="green", highlightbackground="black", highlightthickness=1).grid(row=1, column=1, sticky="nsew")
    
    sem_config_box = Frame(all_config_box, bg="#344E89", highlightbackground="black", highlightthickness=1)
    sem_config_box.grid(row=2, column=0, sticky="nsew")
    if not sem_start:
        Label(sem_config_box, text="Semaines : ", font="Verdana 15", fg="white", bg="#344E89").grid(row=0, column=0, sticky="nsew")
        sem_spin1 = Spinbox(sem_config_box, width=3, font="Verdana 15", from_=1, to=50)
        sem_spin1.grid(row=0, column=1)
        Label(sem_config_box, text=" à ", font="Verdana 15", fg="white", bg="#344E89").grid(row=0, column=2, sticky="nsew")
        sem_spin2 = Spinbox(sem_config_box, width=3, font="Verdana 15", from_=1, to=50)
        sem_spin2.grid(row=0, column=3)
        Button(all_config_box, text="config.", font="Verdana 15", fg="white", bg="#020E26", command=config_sem).grid(row=2, column=1)
    else:
        Label(all_config_box, text="Semaines {} à {}".format(sem_start, sem_end), font="Verdana 15", fg="white", bg="#344E89", highlightbackground="black", highlightthickness=1).grid(row=2, column=0, sticky="nsew")
        Label(all_config_box, text="Fait !", font="Verdana 15", fg="white", bg="green", highlightbackground="black", highlightthickness=1).grid(row=2, column=1, sticky="nsew")

    
    grp_config_box = Frame(all_config_box, bg="#344E89", highlightbackground="black", highlightthickness=1)
    grp_config_box.grid(row=3, column=0, sticky="nsew")
    if not nb_grp:
        Label(grp_config_box, text="Nombre de groupes :", font="Verdana 15", fg="white", bg="#344E89").grid(row=0, column=0)
        grp_spin = Spinbox(grp_config_box, width=3, font="Verdana 15", from_=1, to=50)
        grp_spin.grid(row=0, column=1)
        Button(all_config_box, text="config.", font="Verdana 15", fg="white", bg="#020E26", command=config_grp).grid(row=3, column=1)
    else:
        Label(all_config_box, text="Nombre de groupes : {}".format(nb_grp), font="Verdana 15", fg="white", bg="#344E89", highlightbackground="black", highlightthickness=1).grid(row=3, column=0, sticky="nsew")
        Label(all_config_box, text="Fait !", font="Verdana 15", fg="white", bg="green", highlightbackground="black", highlightthickness=1).grid(row=3, column=1, sticky="nsew")

    if not dic_colleurs:
        Label(all_config_box, text="Colleurs :", font="Verdana 15", fg="white", bg="#344E89", highlightbackground="black", highlightthickness=1).grid(row=4, column=0, sticky="nsew")
        Button(all_config_box, text="config.", font="Verdana 15", fg="white", bg="#020E26", command=config_colleurs).grid(row=4, column=1)
    else:
        Label(all_config_box, text="Colleurs configurés", font="Verdana 15", fg="white", bg="#344E89", highlightbackground="black", highlightthickness=1).grid(row=4, column=0, sticky="nsew")
        Label(all_config_box, text="Fait !", font="Verdana 15", fg="white", bg="green", highlightbackground="black", highlightthickness=1).grid(row=4, column=1, sticky="nsew")

    if not dic_matieres:
        Label(all_config_box, text="Matières :", font="Verdana 15", fg="white", bg="#344E89", highlightbackground="black", highlightthickness=1).grid(row=5, column=0, sticky="nsew")
        Button(all_config_box, text="config.", font="Verdana 15", fg="white", bg="#020E26", command=config_matiere).grid(row=5, column=1)
    else:
        Label(all_config_box, text="Matières configurées", font="Verdana 15", fg="white", bg="#344E89", highlightbackground="black", highlightthickness=1).grid(row=5, column=0, sticky="nsew")
        Label(all_config_box, text="Fait !", font="Verdana 15", fg="white", bg="green", highlightbackground="black", highlightthickness=1).grid(row=5, column=1, sticky="nsew")

    Button(colonne_3, text="Valider", cursor="hand2", font="Verdana 20 bold", fg="white", activeforeground="white", bg="#020E26", activebackground="#020E26", bd=3, command=validate_all).pack(pady=5)
    
    image_back = Image.open("data/img/return_button.png")
    photo_back = ImageTk.PhotoImage(image_back)
    Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=create_or__verif).pack(pady=15)
    
def reload_info():
    """
    ROLE: Interface pour accéder aux informations (action du bouton menu.Information)
    """
    def open_docu(event):
        """
        ROLE: Ouvre la page internet vers le google doc du manuel d'utilisation
        """
        webbrowser.open("https://docs.google.com/document/d/1liZScxKJkO_Epsa_LwyxaCV3nNdL5So7u2nPQAUtuOQ/edit?usp=sharing")
    
    def open_credit(event):
        """
        ROLE: Ouvre la page internet vers le github de fibou
        """
        webbrowser.open("https://github.com/fibouu")
    
    global colonne_3, photo_back
    
    clear(colonne_3)
    
    Label(colonne_3, text="Informations", font="Verdana 25 bold", fg="white", bg="#020E26").pack(pady=15)
    
    info_box = Frame(colonne_3, bg="#020E26")
    info_box.pack(expand=True)
    
    info1_label = Label(info_box, cursor="hand2", font="Verdana 16", bg="#020E26", fg="white", text="L'ensemble de la documentation\nnécessaire au bon fonctionnement du\nprogramme est disponible dans le PDF,\naccessible en cliquant sur ce texte.")
    info1_label.pack()
    info1_label.bind("<Button-1>", open_docu)
    
    info2_label = Label(info_box, font="Verdana 16", bg="#020E26", fg="white", text="L'onglet colloscope (regroupant le créateur\net le vérificateur) fonctionne selon les\nfichiers de colloscope de la plateforme\ne-colle (format et disposition des éléments).")
    info2_label.pack(pady=25)
    
    version_label = Label(colonne_3, text="Version 1.0", font="Verdana 10", bg="#020E26", fg="white")
    version_label.pack()
    
    credit_label = Label(colonne_3, cursor="hand2", text="Développé par Mattéo Demange", font="Verdana 10", bg="#020E26", fg="white")
    credit_label.pack(pady=15)
    credit_label.bind("<Button-1>", open_credit)
    
    image_back = Image.open("data/img/return_button.png")
    photo_back = ImageTk.PhotoImage(image_back)
    Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=back_button_func).pack(pady=15)

def reload_settings():
    """
    ROLE: Interface des paramètres (action du bouton menu.Paramètres)
    STATUS: USELESS
    """
    global colonne_3, photo_back

    clear(colonne_3)
    
    Label(colonne_3, text="Paramètres", font="Verdana 25 bold", fg="white", bg="#020E26").pack(pady=15)
    
    settings_box = Frame(colonne_3, bg="#020E26")
    settings_box.pack(expand=True)
    
    Label(settings_box, text="Aucun paramètre n'a actuellement été implémenté", font="Verdana 12", bg="#020E26", fg="white").pack()
    
    image_back = Image.open("data/img/return_button.png")
    photo_back = ImageTk.PhotoImage(image_back)
    Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=back_button_func).pack(pady=15)

def clear(window):
    """
    ROLE: Supprime l'ensemble des éléments d'une fenêtre
    PARAMTRES: window -> Frame()
    """
    fen.unbind("<MouseWheel>")
    for pack in window.pack_slaves():
        pack.destroy()

def create_or__verif():
    """
    ROLE: Interface pour le choix entre vérificateur ou créateur de khôlloscope (action du bouton menu.Colloscope)
    """
    global photo_back
    
    clear(colonne_3)
    
    Label(colonne_3, text="Colloscope", font="Verdana 25 bold", fg="white", bg="#020E26").pack(pady=15)
    
    button_choice_box = Frame(colonne_3, bg="#020E26")
    button_choice_box.pack(expand=True)

    Button(button_choice_box, text="Créer un colloscope", cursor="hand2", font="Verdana 20 bold", fg="white", activeforeground="white", bg="#020E26", activebackground="#020E26", bd=3, width=19, command=create_colloscope_gui).pack(pady=25)
    Button(button_choice_box, text="Vérifier un colloscope", cursor="hand2", font="Verdana 20 bold", fg="white", activeforeground="white", bg="#020E26", activebackground="#020E26", bd=3, width=19, command=choose_file).pack(pady=35)

    image_back = Image.open("data/img/return_button.png")
    photo_back = ImageTk.PhotoImage(image_back)
    Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=back_button_func).pack(pady=15)

def reload_main():
    """
    ROLE: Interface initiale de l'application (action du bouton menu.Menu)
    """
    global photo1, photo2, photo3, photo4, photo5, photo6, colonne_1_2, colonne_3, ligne1_button, ligne2_button, ligne3_button, ligne4_button, ligne5_button, ligne6_button
    colonne_1_2 = Frame(fen, bg="#344E89")
    colonne_1_2.grid(row=0, column=1, sticky="nsew")

    ligne1 = Frame(colonne_1_2, bg="#344E89", highlightbackground="black", highlightthickness=1)
    ligne1.pack()
    image1 = Image.open("data/img/trois_traits.png")
    photo1 = ImageTk.PhotoImage(image1)
    Button(ligne1, image=photo1, bg="#152C5E", bd=0, cursor="hand2", activebackground="#152C5E", fg="white", relief=FLAT, command=back_button_func).grid(row=0, column=1)
    ligne1_button = Button(ligne1, text="Menu", cursor="hand2", font="Verdana 17 bold", activebackground="#344E89", fg="white", bg="#344E89", relief="flat", bd=0, width=14, command=back_button_func)
    ligne1_button.grid(row=0, column=2)
   
    ligne2 = Frame(colonne_1_2, bg="#344E89", highlightbackground="black", highlightthickness=1)
    ligne2.pack()
    image2 = Image.open("data/img/white_sheets.png")
    photo2 = ImageTk.PhotoImage(image2)
    Button(ligne2, image=photo2, bg="#152C5E", bd=0, cursor="hand2", activebackground="#152C5E", fg="white", relief=FLAT, command=create_or__verif).grid(row=0, column=1)
    ligne2_button = Button(ligne2, text="Colloscope", cursor="hand2", font="Verdana 17 bold", activebackground="#344E89", fg="white", bg="#344E89", relief="flat", bd=0, width=14, command=create_or__verif)
    ligne2_button.grid(row=0, column=2)
    
    ligne3 = Frame(colonne_1_2, bg="#344E89", highlightbackground="black", highlightthickness=1)
    ligne3.pack()
    image3 = Image.open("data/img/dices.png")
    photo3 = ImageTk.PhotoImage(image3)
    Button(ligne3, image=photo3, bg="#152C5E", bd=0, cursor="hand2", activebackground="#152C5E", fg="white", relief=FLAT, command=reload_tas).grid(row=0, column=1)
    ligne3_button = Button(ligne3, text="Tirage au sort", cursor="hand2", font="Verdana 17 bold", activebackground="#344E89", fg="white", bg="#344E89", relief="flat", bd=0, width=14, command=reload_tas)
    ligne3_button.grid(row=0, column=2)
    
    ligne4 = Frame(colonne_1_2, bg="#344E89", highlightbackground="black", highlightthickness=1)
    ligne4.pack()
    image4 = Image.open("data/img/info.png")
    photo4 = ImageTk.PhotoImage(image4)
    Button(ligne4, image=photo4, bg="#152C5E", bd=0, cursor="hand2", activebackground="#152C5E", fg="white", relief=FLAT, command=reload_info).grid(row=0, column=1)
    ligne4_button = Button(ligne4, text="Informations", cursor="hand2", font="Verdana 17 bold", activebackground="#344E89", fg="white", bg="#344E89", relief="flat", bd=0, width=14, command=reload_info)
    ligne4_button.grid(row=0, column=2)

    ligne5 = Frame(colonne_1_2, bg="#344E89", highlightbackground="black", highlightthickness=1)
    ligne5.pack()
    image5 = Image.open("data/img/parametre.png")
    photo5 = ImageTk.PhotoImage(image5)
    Button(ligne5, image=photo5, bg="#152C5E", bd=0, cursor="hand2", activebackground="#152C5E", fg="white", relief=FLAT, command=reload_settings).grid(row=0, column=1)
    ligne5_button = Button(ligne5, text="Paramètres", cursor="hand2", font="Verdana 17 bold", activebackground="#344E89", fg="white", bg="#344E89", relief="flat", bd=0, width=14, command=reload_settings)
    ligne5_button.grid(row=0, column=2)

    ligne6 = Frame(colonne_1_2, bg="#344E89", highlightbackground="black", highlightthickness=1)
    ligne6.pack()
    image6 = Image.open("data/img/eteindre.png")
    photo6 = ImageTk.PhotoImage(image6)
    Button(ligne6, image=photo6, bg="#152C5E", bd=0, cursor="hand2", activebackground="#152C5E", fg="white", relief=FLAT, command=leave_prgm).grid(row=0, column=1)
    ligne6_button = Button(ligne6, text="Quitter", cursor="hand2", font="Verdana 17 bold", activebackground="#344E89", fg="white", bg="#344E89", relief="flat", bd=0, width=14, command=leave_prgm)
    ligne6_button.grid(row=0, column=2)

    ligne7 = Frame(colonne_1_2, bg="#344E89", highlightbackground="black", highlightthickness=1)
    ligne7.pack()
    image7 = Image.open("data/img/empty_logo.png")
    photo7 = ImageTk.PhotoImage(image7)
    Label(ligne7, image=photo7, bg="#152C5E", bd=1, height=200).grid(row=0, column=1)
    ligne7_button = Button(ligne7, font="Verdana 17 bold", bg="#344E89", width=14, height=5, relief="flat", bd=0, activebackground="#344E89")
    ligne7_button.grid(row=0, column=2)

    colonne_3 = Frame(fen, bg="#020E26")
    colonne_3.grid(row=0, column=2, sticky="nsew")

    Label(colonne_3, text="CPGE Helper", bg="#020E26", fg="white", font="Verdana 46 bold").pack(expand=True, anchor="center")

def excel_to_list(file_path):
    """
    ROLE: Conversion d'un fichier XLSX en une liste pratique
    PARAMETRE: file_path -> str
    RETURN: list
    STATUS: fonction à refaire pour la rendre fonctionnelle avec e-colle
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active
    excel_data = []
    max_row_length = 0

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_data = []
        for cell in row:
            if isinstance(cell.value, datetime):
                row_data.append(cell.value.strftime("%d/%m/%Y"))
            elif cell.value is not None:
                row_data.append(str(cell.value))
            else:
                row_data.append("None")
        
        max_row_length = max(max_row_length, len(row_data))
        excel_data.append(row_data)

    # Remplir les cases vides avec la chaîne "None"
    for row_data in excel_data:
        row_data.extend(["None"] * (max_row_length - len(row_data)))

    wb.close()
    return excel_data

def choose_file():
    """
    ROLE: Interface pour choisir un fichier XLSX à verifier (action du bouton Verififier un colloscope)
    STATUS: fonction à refaire pour la rendre fonctionnelle avec e-colle
    """
    global canvas, error_box, photo_back
    def recap_grp(tab):
        def check_pattern_mipcal(chaine): #MathInfoPhysChimAngLett
            return compile(r'^(P|L|Ang|C|M|I)\d+').match(chaine)
        
        def look_for_sem(tab):
            for l in range(l_len):
                if tab[l][0].lower().startswith("semaine"):
                    return l
        
        grp_colle = {} #dico {NB_GRP: ((coordX, coordY), (heure, date))}
        l_len = len(tab) #nombre de lignes
        c_len = len(tab[0]) #nombre de colonnes
        sem_line = look_for_sem(tab)
        for l in range(l_len):
            if check_pattern_mipcal(tab[l][0]):
                hours = tab[l][1].rstrip().rsplit(' ', 1)
                if (len(hours)) > 1:
                    hours = hours[-1].rsplit('-', 1)
                    for i in range(2):
                        hours[i] = int(hours[i])
                for c in range(c_len):
                    if tab[l][c].isdigit():
                        if int(tab[l][c]) in grp_colle:
                            grp_colle[int(tab[l][c])].append(((l, c), (tab[sem_line][c], hours, tab[l][1][:3])))
                        else:
                            grp_colle[int(tab[l][c])] = [((l, c), (tab[sem_line][c], hours, tab[l][1][:3]))]
        return grp_colle
    
    def find_error(dico):
        """
        ROLE: Compare tout les éléments pour trouver ceux incohérents
        PARAMETRE: dico -> dic
        RETURN: lst
        """
        ans = []
        for key1, value in dico.items():
            for i in range(len(value)):
                for j in range(len(value)):
                    if not i == j and value[i][1][0] == value[j][1][0]:
                        if value[i][1] == (value[j][1][0], [value[j][1][1][0]-1, value[j][1][1][1]-1], value[j][1][2]):
                            ans.append([(value[i][1], value[j][1]), key1])
        return ans
    
    def error_gui(file_name):
        """
        ROLE: Interface pour l'affichage des incohérences dans le khôlloscope sélectionné
        PARAMETRE: file_name -> str
        """
        def on_frame_configure(event):
            """
            ROLE: Permet de scroller sur toute  la zone de canvas
            """
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        global canvas, error_box, photo_back
        clear(colonne_3)
        
        l = excel_to_list(file_path) 
        colloscope = recap_grp(l)
        
        error_list = find_error(colloscope)
        
        Label(colonne_3, text="Vérificateur de colloscope", font="Verdana 22 bold", fg="white", bg="#020D26").pack(pady=15)
        middle_data_box = Frame(colonne_3, bg="#020E26")
        middle_data_box.pack(expand=True)
        
        file_frame = Frame(middle_data_box, bg="#020E26")
        file_frame.pack()
        Label(file_frame, text="Fichier :", font="Verdana 15 bold", fg="white", bg="#020D26").grid(row=0, column=1)
        if len(file_name) > 29:
            file_name = file_name[:26] + "..."
        Button(file_frame, text=(file_name), cursor="hand2", font="Verdana 15 bold", fg="#324F87", bg="#020D26", activebackground="#020D26", activeforeground="#324F87", relief="sunken", bd=0, command=choose_file).grid(row=0, column=2)
        
        error_box_frame = Frame(middle_data_box, bg="#344E89")
        error_box_frame.pack(pady=5)
        
        canvas = Canvas(error_box_frame, width=350, height=200, background="#344E89")
        error_box = Frame(canvas, background="#344E89")
        vsb = Scrollbar(error_box_frame, orient=VERTICAL, bg="#344E89", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        fen.bind("<MouseWheel>", on_mousewheel)
        
        vsb.pack(side=RIGHT, fill=Y)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        canvas.create_window((4, 4), window=error_box, anchor=NW, tags="error_box")

        if error_list:
            for i, el in enumerate(error_list):
                Label(error_box, text="Groupe {}".format(el[1]), font="Verdana 10 bold", bg="#344E89", fg="white").grid(row=i, column=0, padx=5, pady=5)
                Label(error_box, text="Semaine {}".format(el[0][0][0]), font="Verdana 10 bold", bg="#344E89", fg="white").grid(row=i, column=1, padx=5, pady=5)
                Label(error_box, text="{}h-{}h".format(el[0][0][1][0], el[0][0][1][1]), font="Verdana 10 bold", bg="#344E89", fg="white").grid(row=i, column=2, padx=5, pady=5)
                Label(error_box, text="{}h-{}h".format(el[0][1][1][0], el[0][1][1][1]), font="Verdana 10 bold", bg="#344E89", fg="white").grid(row=i, column=3, padx=5, pady=5)
        else:
            Label(error_box, text="Aucune erreur détectée", font="Verdana 16 bold", bg="#344E89", fg="white").grid(row=0, column=0, padx=5, pady=5)

        canvas.bind("<Configure>", on_frame_configure)
        
        image_back = Image.open("data/img/return_button.png")
        photo_back = ImageTk.PhotoImage(image_back)
        Button(colonne_3, image=photo_back, cursor="hand2", bg="#020E26", activebackground="#020E26", bd=0, command=back_button_func).pack(pady=15)
    
    clear(colonne_3)
    
    file_path = filedialog.askopenfilename(title="Ouvrir un fichier", filetypes=(("Fichiers Excel", "*.xlsx"), ("XLSX", "*.xlsx")))
    if file_path == "":
        create_or__verif()
    else:
        error_gui(basename(file_path))

#######################
##PROGRAMME PRINCIPAL##
#######################

fen = Tk()
fen.geometry("800x500") #750x425
fen.resizable(width=False, height=False)
fen.title("CPGE Helper")
fen.wm_iconbitmap("data/img/Logo-CPGE.ico")
fen.configure(bg="#020D26")
fen.grid_rowconfigure(0, weight=1)
fen.grid_columnconfigure(2, weight=1)

photo1, photo2, photo3, photo4, photo5, photo6, photo_back, photo_add = None, None, None, None, None, None, None, None
reload_main()

fen.mainloop()